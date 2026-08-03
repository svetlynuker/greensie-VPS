"""PPA nabídka jako Excel s **živými vzorci** – model k ručnímu doladění.

Proti PDF (nabídka pro zákazníka, whitelist zákaznických polí) je tohle
**interní výpočtový model**: CAPEX, marže, provize, IRR i zisk SPV. Vzniká
z uloženého výsledku výpočtu (`navrhovana_reseni.popis_json`) a z parametrů,
se kterými se počítalo.

**Do sešitu se nesypou hotová čísla, ale vzorce.** Kdyby to byl jen snímek
hodnot, nešlo by v něm nic doladit – a přesně kvůli tomu vzniká. Přepíšeš
žlutou buňku (cenu PPA, marži, servis) a přepočítá se DSCR, IRR i odkupní
tabulka, stejně jako v původním `docs/PPA výpočet.xlsx`.

Rozvržení je proti tomu Excelu srovnané do pěti listů:

* **Zadání** – všechny vstupy na jednom místě (žlutě) a pod nimi odvozená
  investice. V původním souboru byly vstupy rozeseté po listech (C7 podíl
  samospotřeby, F6 cena, B13 splátka, E32 marže…) a hledaly se.
* **Cashflow** – roky ve sloupcích jako dřív, ale DSCR pod limitem se obarví
  a délka jde podle skutečného kontraktu, ne napevno 15 nebo 10 let.
* **Odkup** – odkupní tabulka samostatně, ne zamíchaná do sloupců V–AI.
* **Splátkový kalendář** – dopočítaný vzorci. V původním souboru byl nakliknutý
  natvrdo, takže po změně úroku ukazoval čísla ze starého zadání.
* **Úspora zákazníka** – co z toho má klient. V původním Excelu chybělo.

Buňky se nezamykají. Žlutá znamená „tohle je vstup", ale doladit jde cokoli –
zámek by při ručním hledání varianty jen překážel.

Vzorce musí odpovídat funkcím v `ppa_v2` řádek po řádku. Když se změní výpočet
(indexace, odkup, nájem baterie), musí se změnit i tady, jinak sešit tiše
ukáže jiná čísla než appka. Hlídá to `tests/test_excel_ppa.py`.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from app.nabidkovac.ppa_v2 import ParametryEkonomiky

MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PRIPONA = ".xlsx"

# ---------------------------------------------------------------- vzhled
ZLUTA = PatternFill("solid", fgColor="FFF2CC")  # editovatelný vstup
SEDA = PatternFill("solid", fgColor="EDEDED")  # hlavička sekce
CERVENA = PatternFill("solid", fgColor="FFC7CE")  # DSCR pod limitem
NADPIS = Font(bold=True, size=14)
SEKCE = Font(bold=True, size=11)
TUCNE = Font(bold=True)
SEDY_TEXT = Font(size=9, color="808080")
_tenka = Side(style="thin", color="D0D0D0")
RAMECEK = Border(left=_tenka, right=_tenka, top=_tenka, bottom=_tenka)

# Formáty se zapisují v anglické notaci, Excel je zobrazí podle lokalizace –
# v českém Excelu tedy s mezerou mezi tisíci a desetinnou čárkou.
KC = '#,##0 "Kč"'
KC_PRESNE = '#,##0.00 "Kč"'
MWH = '#,##0.0'
PROCENTO = '0.0%'
PROCENTO_PRESNE = '0.00%'
NASOBEK = '0.000'
CISLO = '#,##0.0'


def _popis(ws, radek: int, text: str, jednotka: str = "", poznamka: str = "") -> None:
    """Popisek řádku v zadání (sloupec B), jednotka (D) a poznámka (E)."""
    ws.cell(row=radek, column=2, value=text)
    if jednotka:
        ws.cell(row=radek, column=4, value=jednotka)
    if poznamka:
        bunka = ws.cell(row=radek, column=5, value=poznamka)
        bunka.font = SEDY_TEXT


def _vstup(ws, radek: int, hodnota, format_cisla: str) -> str:
    """Žlutá editovatelná buňka v zadání (sloupec C). Vrací svou adresu."""
    bunka = ws.cell(row=radek, column=3, value=hodnota)
    bunka.fill = ZLUTA
    bunka.border = RAMECEK
    bunka.number_format = format_cisla
    return f"$C${radek}"


def _odvozeno(ws, radek: int, vzorec: str, format_cisla: str) -> str:
    """Bílá počítaná buňka v zadání (sloupec C). Vrací svou adresu."""
    bunka = ws.cell(row=radek, column=3, value=vzorec)
    bunka.number_format = format_cisla
    return f"$C${radek}"


def _sekce(ws, radek: int, text: str) -> None:
    for sloupec in range(2, 6):
        ws.cell(row=radek, column=sloupec).fill = SEDA
    bunka = ws.cell(row=radek, column=2, value=text)
    bunka.font = SEKCE


def _hlavicka_tabulky(ws, radek: int, nazvy: list[str], od_sloupce: int = 2) -> None:
    for i, nazev in enumerate(nazvy):
        bunka = ws.cell(row=radek, column=od_sloupce + i, value=nazev)
        bunka.font = TUCNE
        bunka.fill = SEDA
        bunka.border = RAMECEK
        bunka.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")


def _sirky(ws, sirky: dict[str, float]) -> None:
    for sloupec, sirka in sirky.items():
        ws.column_dimensions[sloupec].width = sirka


def _podil(varianta: dict, klic_mwh: str, klic_podilu: str) -> float:
    """Podíl z výroby dopočítaný z MWh prvního roku, ne z uloženého procenta.

    `energie.mira_samospotreby` je v `popis_json` zaokrouhlená na čtyři desetinná
    místa; kdyby se použila přímo, sešit by ukazoval o desítky tisíc jiné tržby
    než appka a vypadalo by to jako chyba modelu.
    """
    rok1 = (varianta.get("roky_investor") or [{}])[0]
    vyroba = float(rok1.get("vyroba_mwh") or 0)
    if vyroba > 0 and rok1.get(klic_mwh) is not None:
        return float(rok1[klic_mwh]) / vyroba
    return float((varianta.get("energie") or {}).get(klic_podilu) or 0)


def _nazev(wb: Workbook, jmeno: str, list_nazev: str, adresa: str) -> None:
    """Pojmenovaná buňka, aby vzorce byly čitelné (`=cena_ppa`, ne `=Zadání!$C$27`)."""
    wb.defined_names[jmeno] = DefinedName(jmeno, attr_text=f"'{list_nazev}'!{adresa}")


# ---------------------------------------------------------------- list Zadání
def _list_zadani(
    wb: Workbook,
    hlavicka: dict,
    varianta: dict,
    vstup: dict,
    p: ParametryEkonomiky,
) -> None:
    """Vstupy (žlutě) a z nich odvozená investice (vzorci)."""
    ws = wb.active
    ws.title = "Zadání"
    _sirky(ws, {"A": 2, "B": 34, "C": 16, "D": 10, "E": 62})

    ws["B1"] = hlavicka["titulek"]
    ws["B1"].font = NADPIS
    ws["B2"] = hlavicka["podtitulek"]
    ws["B2"].font = SEDY_TEXT
    ws["B3"] = hlavicka["vygenerovano"]
    ws["B3"].font = SEDY_TEXT

    baterie = varianta.get("baterie") or {}
    energie = varianta.get("energie") or {}
    fin = varianta.get("financovani") or {}
    delka = int(varianta.get("delka_kontraktu_roky") or 0)

    # Nákladová cena FVE se z uloženého výsledku nedá oddělit od baterie (obojí je
    # sečtené v `financovani`), takže se dopočítá z kWp a ceny za kWp – přesně tak,
    # jak ji poskládal `spocti_variantu`.
    r = 5
    _sekce(ws, r, "Elektrárna")
    r += 1
    _popis(ws, r, "Instalovaný výkon", "kWp", "z návrhu na cíl samospotřeby")
    a_kwp = _vstup(ws, r, float(varianta.get("kwp") or 0), CISLO)
    r += 1
    _popis(ws, r, "Nákladová cena", "Kč/kWp", "nákup technologie")
    a_cena_kwp = _vstup(ws, r, p.nakladova_cena_kc_kwp, KC)
    r += 1
    _popis(ws, r, "Marže FVE", "×", "prodejní cena do SPV je nákladová × marže")
    a_marze_fve = _vstup(ws, r, p.marze_fve, NASOBEK)
    r += 1
    _popis(ws, r, "Provize FVE", "%", "z prodejní ceny")
    a_provize_fve = _vstup(ws, r, p.provize_fve, PROCENTO)

    r += 2
    _sekce(ws, r, "Baterie")
    r += 1
    _popis(ws, r, "Kapacita", "kWh", "0 = nabídka bez baterie")
    a_bat_kwh = _vstup(ws, r, float(baterie.get("kapacita_kwh") or 0), CISLO)
    r += 1
    _popis(ws, r, "Výkon", "kW")
    _vstup(ws, r, float(baterie.get("vykon_kw") or 0), CISLO)
    r += 1
    _popis(ws, r, "Nákladová cena baterie", "Kč")
    a_bat_naklad = _vstup(ws, r, float(baterie.get("nakladova_cena_kc") or 0), KC)
    r += 1
    _popis(ws, r, "Marže BESS", "×")
    a_marze_bess = _vstup(ws, r, p.marze_bess, NASOBEK)
    r += 1
    _popis(ws, r, "Provize BESS", "%")
    a_provize_bess = _vstup(ws, r, p.provize_bess, PROCENTO)
    r += 1
    _popis(ws, r, "Marže v nájmu baterie", "Kč/měs", "nájem je marže + splátka + EMS")
    a_bess_marze = _vstup(ws, r, p.bess_marze_kc_mesic, KC)
    r += 1
    _popis(ws, r, "EMS", "Kč/měs")
    a_ems = _vstup(ws, r, p.bess_ems_kc_mesic, KC)
    r += 1
    _popis(ws, r, "Servis baterie", "Kč/rok")
    a_bess_servis = _vstup(ws, r, p.bess_servis_kc_rok, KC)

    r += 2
    _sekce(ws, r, "Energie")
    r += 1
    _popis(ws, r, "Výroba v roce 1", "MWh", "simulace z lokality, sklonu a azimutu")
    a_vyroba1 = _vstup(ws, r, float(energie.get("vyroba_rok1_mwh") or 0), MWH)
    r += 1
    _popis(
        ws, r, "Podíl samospotřeby", "%",
        "z výroby – spárováno s 15min profilem klienta (v původním Excelu ručních 78 %)",
    )
    a_podil_ss = _vstup(ws, r, _podil(varianta, "samospotreba_mwh", "mira_samospotreby"), PROCENTO)
    r += 1
    _popis(ws, r, "Podíl zpeněženého přebytku", "%", "z výroby – export / sdílení")
    a_podil_sdil = _vstup(ws, r, _podil(varianta, "sdileni_mwh", "podil_exportu"), PROCENTO)
    r += 1
    _popis(ws, r, "Roční degradace výroby", "%")
    a_degradace = _vstup(ws, r, p.degradace_rocni, PROCENTO_PRESNE)

    r += 2
    _sekce(ws, r, "Ceny")
    r += 1
    _popis(
        ws, r, "Cena PPA v roce 1", "Kč/MWh",
        "spočtená nejnižší cena, která projde bankou i investorem – tohle laď jako první",
    )
    a_cena_ppa = _vstup(ws, r, float(varianta.get("cena_ppa_kc_mwh") or 0), KC)
    r += 1
    _popis(ws, r, "Cena za export / sdílení", "Kč/MWh", "0 = za přetoky se neinkasuje nic")
    a_cena_export = _vstup(ws, r, float(varianta.get("cena_exportu_kc_mwh") or 0), KC)
    r += 1
    _popis(ws, r, "Indexace – krok", "%", "cena drží periodu plochá a pak skočí")
    a_index_krok = _vstup(ws, r, p.indexace_krok, PROCENTO)
    r += 1
    _popis(ws, r, "Indexace – perioda", "let")
    a_index_perioda = _vstup(ws, r, int(p.indexace_perioda_roky), "0")
    r += 1
    _popis(
        ws, r, "Vyhnutelná cena zákazníka", "Kč/MWh",
        "silová složka + vyhnutelné regulované – proti tomu se počítá úspora",
    )
    a_cena_vyhnutelna = _vstup(ws, r, float(varianta.get("cena_vyhnutelna_kc_mwh") or 0), KC)

    r += 2
    _sekce(ws, r, "Provoz a financování")
    r += 1
    _popis(ws, r, "Servis FVE", "Kč/rok")
    a_servis = _vstup(ws, r, p.servis_kc_rok, KC)
    r += 1
    _popis(ws, r, "Délka kontraktu a úvěru", "let")
    a_roky = _vstup(ws, r, delka, "0")
    r += 1
    _popis(ws, r, "Podíl vlastního kapitálu", "%", "zbytek je bankovní úvěr")
    a_podil_vk = _vstup(ws, r, p.podil_vlastniho_kapitalu, PROCENTO)
    r += 1
    _popis(ws, r, "Úroková sazba", "% p.a.")
    a_urok = _vstup(ws, r, p.urokova_sazba, PROCENTO_PRESNE)
    r += 1
    _popis(ws, r, "DSCR minimum", "×", "bankovní kovenant – pod tím se rok obarví červeně")
    a_dscr_min = _vstup(ws, r, p.dscr_min, NASOBEK)
    r += 1
    _popis(ws, r, "Cílové IRR investora", "%")
    a_irr_cil = _vstup(ws, r, p.irr_cil, PROCENTO)
    r += 1
    _popis(ws, r, "Diskontní sazba pro NPV", "%")
    a_diskont = _vstup(ws, r, p.diskontni_sazba, PROCENTO_PRESNE)

    r += 2
    _sekce(ws, r, "Odkup technologie")
    r += 1
    _popis(ws, r, "Roční poplatek", "%", "ze zůstatku fiktivního úvěru na 100 % CAPEX")
    a_odkup_poplatek = _vstup(ws, r, p.odkup_poplatek_rocni, PROCENTO_PRESNE)
    r += 1
    _popis(ws, r, "Poplatek za předčasné splacení", "%", "z reálného zůstatku úvěru")
    a_odkup_predcasne = _vstup(ws, r, p.odkup_poplatek_predcasne_splaceni, PROCENTO)

    # ---- odvozené (bílé, vzorce)
    r += 2
    _sekce(ws, r, "Investice (dopočítané ze zadání výše)")
    r += 1
    _popis(ws, r, "Nákladová cena FVE", "Kč", "výkon × cena za kWp")
    a_naklad_fve = _odvozeno(ws, r, "=kwp*cena_kwp", KC)
    r += 1
    _popis(ws, r, "CAPEX FVE (prodej do SPV)", "Kč", "nákladová × marže")
    a_capex_fve = _odvozeno(ws, r, "=naklad_fve*marze_fve", KC)
    r += 1
    _popis(ws, r, "Nákladová cena baterie", "Kč")
    a_naklad_bess = _odvozeno(ws, r, "=bat_naklad", KC)
    r += 1
    _popis(ws, r, "CAPEX baterie", "Kč")
    a_capex_bess = _odvozeno(ws, r, "=naklad_bess*marze_bess", KC)
    r += 1
    _popis(ws, r, "CAPEX celkem", "Kč", "to, co se financuje")
    a_capex = _odvozeno(ws, r, "=capex_fve+capex_bess", KC)
    r += 1
    _popis(ws, r, "Provize celkem", "Kč")
    a_provize = _odvozeno(ws, r, "=capex_fve*provize_fve+capex_bess*provize_bess", KC)
    r += 1
    _popis(ws, r, "Zisk Greensie hned", "Kč", "CAPEX − nákladová − provize")
    _odvozeno(ws, r, "=capex-naklad_fve-naklad_bess-provize", KC)
    ws.cell(row=r, column=3).font = TUCNE
    r += 1
    _popis(ws, r, "Vlastní kapitál", "Kč")
    a_vk = _odvozeno(ws, r, "=capex*podil_vk", KC)
    r += 1
    _popis(ws, r, "Bankovní úvěr", "Kč")
    a_uver = _odvozeno(ws, r, "=capex-vlastni_kapital", KC)
    r += 1
    _popis(ws, r, "Měsíční splátka úvěru", "Kč", "anuita")
    a_splatka_mesic = _odvozeno(
        ws, r, "=IF(OR(uver<=0,roky<=0),0,-PMT(urok/12,roky*12,uver))", KC_PRESNE
    )
    r += 1
    _popis(ws, r, "Roční splátka úvěru", "Kč")
    a_splatka_rok = _odvozeno(ws, r, "=splatka_mesic*12", KC)
    r += 1
    _popis(ws, r, "Splátka úvěru na baterii", "Kč/měs", "část anuity připadající na BESS")
    a_splatka_bess = _odvozeno(
        ws,
        r,
        "=IF(OR(capex_bess<=0,roky<=0),0,-PMT(urok/12,roky*12,capex_bess*(1-podil_vk)))",
        KC_PRESNE,
    )
    r += 1
    _popis(ws, r, "Nájem baterie", "Kč/měs", "marže + splátka baterie + EMS")
    a_najem_mesic = _odvozeno(
        ws, r, "=IF(capex_bess<=0,0,bess_marze_mesic+splatka_bess+ems_mesic)", KC_PRESNE
    )
    r += 1
    _popis(ws, r, "Nájem baterie", "Kč/rok")
    a_najem_rok = _odvozeno(ws, r, "=najem_mesic*12", KC)
    r += 1
    _popis(ws, r, "Provozní náklady", "Kč/rok", "servis FVE + servis baterie + EMS")
    a_naklady_rok = _odvozeno(
        ws, r, "=servis_rok+IF(capex_bess<=0,0,bess_servis_rok+ems_mesic*12)", KC
    )

    for jmeno, adresa in {
        "kwp": a_kwp,
        "cena_kwp": a_cena_kwp,
        "marze_fve": a_marze_fve,
        "provize_fve": a_provize_fve,
        "bat_kwh": a_bat_kwh,
        "bat_naklad": a_bat_naklad,
        "marze_bess": a_marze_bess,
        "provize_bess": a_provize_bess,
        "bess_marze_mesic": a_bess_marze,
        "ems_mesic": a_ems,
        "bess_servis_rok": a_bess_servis,
        "vyroba_rok1": a_vyroba1,
        "podil_ss": a_podil_ss,
        "podil_sdil": a_podil_sdil,
        "degradace": a_degradace,
        "cena_ppa": a_cena_ppa,
        "cena_export": a_cena_export,
        "index_krok": a_index_krok,
        "index_perioda": a_index_perioda,
        "cena_vyhnutelna": a_cena_vyhnutelna,
        "servis_rok": a_servis,
        "roky": a_roky,
        "podil_vk": a_podil_vk,
        "urok": a_urok,
        "dscr_limit": a_dscr_min,
        "irr_cil": a_irr_cil,
        "diskont": a_diskont,
        "odkup_poplatek": a_odkup_poplatek,
        "odkup_predcasne": a_odkup_predcasne,
        "naklad_fve": a_naklad_fve,
        "capex_fve": a_capex_fve,
        "naklad_bess": a_naklad_bess,
        "capex_bess": a_capex_bess,
        "capex": a_capex,
        "provize": a_provize,
        "vlastni_kapital": a_vk,
        "uver": a_uver,
        "splatka_mesic": a_splatka_mesic,
        "splatka_rok": a_splatka_rok,
        "splatka_bess": a_splatka_bess,
        "najem_mesic": a_najem_mesic,
        "najem_rok": a_najem_rok,
        "naklady_rok": a_naklady_rok,
    }.items():
        _nazev(wb, jmeno, "Zadání", adresa)

    r += 2
    ws.cell(
        row=r,
        column=2,
        value=(
            "Žluté buňky jsou vstupy – přepiš je a všechno ostatní se přepočítá. "
            "Bílé buňky jsou vzorce, do těch se sahá jen když měníš samotný model."
        ),
    ).font = SEDY_TEXT
    ws.freeze_panes = "B5"


# ---------------------------------------------------------------- list Cashflow
# Řádky cashflow: (klíč, popisek, vzorec, formát, tučně). Klíč `None` je mezinadpis.
# Ve vzorci `{sloupec}` doplní písmeno roku a `{klíč}` adresu jiného řádku ve stejném
# sloupci – čísla řádků se nikde nepíšou ručně. Když se sem přidá řádek, ostatní
# vzorce se posunou samy; s natvrdo psanými čísly by se tiše rozešly.
RADKY_CASHFLOW: list[tuple[str | None, str, str, str, bool]] = [
    (None, "ENERGIE", "", "", False),
    ("vyroba", "Výroba", "=vyroba_rok1*(1-degradace)^({sloupec}$4-1)", MWH, False),
    ("samospotreba", "Samospotřeba klienta", "={vyroba}*podil_ss", MWH, False),
    ("export", "Export / sdílení", "={vyroba}*podil_sdil", MWH, False),
    (None, "CENY", "", "", False),
    (
        "cena_ppa_rok",
        "Cena PPA",
        "=cena_ppa*(1+index_krok)^INT(({sloupec}$4-1)/index_perioda)",
        KC,
        False,
    ),
    (
        "cena_export_rok",
        "Cena za export",
        "=cena_export*(1+index_krok)^INT(({sloupec}$4-1)/index_perioda)",
        KC,
        False,
    ),
    (None, "VÝNOSY", "", "", False),
    ("prodej_klient", "Prodej klientovi", "={samospotreba}*{cena_ppa_rok}", KC, False),
    ("prodej_export", "Prodej přebytku", "={export}*{cena_export_rok}", KC, False),
    ("najem", "Nájem baterie", "=najem_rok", KC, False),
    (
        "vynosy",
        "Výnosy celkem",
        "={prodej_klient}+{prodej_export}+{najem}",
        KC,
        True,
    ),
    (None, "NÁKLADY A DLUH", "", "", False),
    ("naklady", "Provozní náklady", "=naklady_rok", KC, False),
    ("zdroje", "Provozní zisk (zdroje)", "={vynosy}-{naklady}", KC, True),
    ("splatka", "Splátka úvěru", "=splatka_rok", KC, False),
    ("dscr", "DSCR", '=IF({splatka}=0,"",{zdroje}/{splatka})', NASOBEK, True),
    ("zisk", "Zisk po splátkách", "={zdroje}-{splatka}", KC, True),
]


def _list_cashflow(
    wb: Workbook, delka: int, prvni_rok: int, s_baterii: bool
) -> dict[str, int]:
    """Cashflow investora po letech – roky ve sloupcích, vše vzorci ze Zadání.

    Vrací mapu klíč → číslo řádku, aby na ni šlo odkázat z dalších listů.
    """
    ws = wb.create_sheet("Cashflow")
    _sirky(ws, {"A": 2, "B": 30, "C": 14})
    for i in range(delka):
        ws.column_dimensions[get_column_letter(4 + i)].width = 13

    ws["B1"] = "Cashflow investora (SPV)"
    ws["B1"].font = NADPIS
    ws["B2"] = "Vše počítá ze Zadání. Roky ve sloupcích, sloupec C je rok 0 (investice)."
    ws["B2"].font = SEDY_TEXT

    prvni = 4  # první sloupec s rokem kontraktu
    posledni = prvni + delka - 1

    ws["B4"] = "Rok kontraktu"
    ws["B4"].font = TUCNE
    ws["B5"] = "Kalendářní rok"
    ws["B5"].font = TUCNE
    ws["C4"] = 0
    ws["C4"].font = TUCNE
    ws["C4"].fill = SEDA
    ws["C5"] = prvni_rok - 1
    ws["C5"].fill = SEDA
    for i in range(delka):
        sl = get_column_letter(prvni + i)
        b = ws[f"{sl}4"]
        b.value = i + 1
        b.font = TUCNE
        b.fill = SEDA
        b.alignment = Alignment(horizontal="center")
        ws[f"{sl}5"] = prvni_rok + i
        ws[f"{sl}5"].alignment = Alignment(horizontal="center")

    prvni_radek = 6
    radky = {
        klic: prvni_radek + i
        for i, (klic, *_) in enumerate(RADKY_CASHFLOW)
        if klic is not None
    }

    for i, (klic, popisek, vzorec, format_cisla, tucne) in enumerate(RADKY_CASHFLOW):
        r = prvni_radek + i
        bunka = ws.cell(row=r, column=2, value=popisek)
        if klic is None:
            bunka.font = SEKCE
            for sloupec in range(2, posledni + 1):
                ws.cell(row=r, column=sloupec).fill = SEDA
            continue
        if tucne:
            bunka.font = TUCNE
        # Nájem baterie u nabídky bez baterie zůstává jako nulový řádek, ať je
        # vidět, že se s ním počítalo – jen se zešedí.
        if klic == "najem" and not s_baterii:
            bunka.font = SEDY_TEXT
        for j in range(delka):
            sl = get_column_letter(prvni + j)
            adresy = {jmeno: f"{sl}{radek}" for jmeno, radek in radky.items()}
            c = ws.cell(row=r, column=prvni + j, value=vzorec.format(sloupec=sl, **adresy))
            c.number_format = format_cisla
            if tucne:
                c.font = TUCNE

    # DSCR pod bankovním limitem červeně – v původním Excelu se muselo hlídat okem.
    radek_dscr = radky["dscr"]
    ws.conditional_formatting.add(
        f"{get_column_letter(prvni)}{radek_dscr}:{get_column_letter(posledni)}{radek_dscr}",
        CellIsRule(operator="lessThan", formula=["dscr_limit"], fill=CERVENA),
    )

    r = prvni_radek + len(RADKY_CASHFLOW) + 1
    ws.cell(row=r, column=2, value="VÝSLEDEK INVESTORA").font = SEKCE
    for sloupec in range(2, posledni + 1):
        ws.cell(row=r, column=sloupec).fill = SEDA

    r += 1
    ws.cell(row=r, column=2, value="CF vlastního kapitálu")
    c0 = ws.cell(row=r, column=3, value="=-vlastni_kapital")
    c0.number_format = KC
    for j in range(delka):
        sl = get_column_letter(prvni + j)
        c = ws.cell(row=r, column=prvni + j, value=f"={sl}{radky['zisk']}")
        c.number_format = KC
    radek_cf = r

    cf_prvni = f"{get_column_letter(prvni)}{radek_cf}"
    cf_posledni = f"{get_column_letter(posledni)}{radek_cf}"
    dscr_rozsah = f"{get_column_letter(prvni)}{radek_dscr}:{get_column_letter(posledni)}{radek_dscr}"
    for popisek, vzorec, format_cisla in [
        ("IRR vlastních zdrojů", f"=IRR(C{radek_cf}:{cf_posledni})", PROCENTO_PRESNE),
        ("Cílové IRR", "=irr_cil", PROCENTO_PRESNE),
        ("NPV", f"=C{radek_cf}+NPV(diskont,{cf_prvni}:{cf_posledni})", KC),
        ("DSCR minimum za kontrakt", f"=MIN({dscr_rozsah})", NASOBEK),
        ("DSCR limit banky", "=dscr_limit", NASOBEK),
    ]:
        r += 1
        ws.cell(row=r, column=2, value=popisek).font = TUCNE
        c = ws.cell(row=r, column=3, value=vzorec)
        c.number_format = format_cisla
        c.font = TUCNE

    ws.freeze_panes = "D6"
    return radky


# ---------------------------------------------------------------- list Odkup
def _list_odkup(wb: Workbook, delka: int, prvni_rok: int) -> None:
    """Za kolik si klient technologii odkoupí v roce t a co na tom SPV zůstane."""
    ws = wb.create_sheet("Odkup")
    _sirky(ws, {"A": 2, "B": 8, "C": 12, "D": 20, "E": 18, "F": 18, "G": 18, "H": 18, "I": 16})

    ws["B1"] = "Odkupní tabulka"
    ws["B1"].font = NADPIS
    ws["B2"] = (
        "Odkupní cena se počítá z fiktivního úvěru na 100 % CAPEX – klient platí "
        "zbývající hodnotu celé technologie, ne jen zbytek našeho úvěru."
    )
    ws["B2"].font = SEDY_TEXT

    _hlavicka_tabulky(
        ws,
        4,
        [
            "Rok",
            "Kalendářní rok",
            "Zůstatek fiktivního úvěru (100 % CAPEX)",
            "Kumulativní poplatek",
            "Odkupní cena",
            "Reálný zůstatek úvěru",
            "Poplatek za předčasné splacení",
            "Zisk SPV",
        ],
    )

    # Anuita fiktivního úvěru na celý CAPEX a poplatek za první rok – ten se pak
    # jen násobí pořadím roku (stejně jako `odkupni_tabulka` v ppa_v2).
    ws["K4"] = "pomocné výpočty"
    ws["K4"].font = SEDY_TEXT
    ws["K5"] = "anuita ze 100 % CAPEX"
    ws["K5"].font = SEDY_TEXT
    ws["L5"] = "=IF(OR(capex<=0,roky<=0),0,-PMT(urok/12,roky*12,capex))"
    ws["L5"].number_format = KC_PRESNE
    ws["K6"] = "poplatek za 1. rok"
    ws["K6"].font = SEDY_TEXT
    ws["L6"] = (
        "=IF(capex<=0,0,(capex*(1+urok/12)^12-$L$5*((1+urok/12)^12-1)/(urok/12))*odkup_poplatek)"
    )
    ws["L6"].number_format = KC_PRESNE

    for i in range(delka):
        r = 5 + i
        m = f"MIN($B{r}*12,roky*12)"  # v posledním roce se nesmí přestřelit počet splátek
        ws.cell(row=r, column=2, value=i + 1)
        ws.cell(row=r, column=3, value=prvni_rok + i)
        hodnoty = {
            4: f"=MAX(0,capex*(1+urok/12)^{m}-$L$5*((1+urok/12)^{m}-1)/(urok/12))",
            5: f"=$L$6*$B{r}",
            6: f"=$D{r}+$E{r}",
            7: f"=MAX(0,uver*(1+urok/12)^{m}-splatka_mesic*((1+urok/12)^{m}-1)/(urok/12))",
            8: f"=$G{r}*odkup_predcasne",
            9: f"=$F{r}-$G{r}-$H{r}",
        }
        for sloupec, vzorec in hodnoty.items():
            c = ws.cell(row=r, column=sloupec, value=vzorec)
            c.number_format = KC
            c.border = RAMECEK
        ws.cell(row=r, column=6).font = TUCNE
        ws.cell(row=r, column=9).font = TUCNE

    ws.freeze_panes = "B5"


# ---------------------------------------------------------------- list Splátkový kalendář
def _list_splatky(wb: Workbook, delka: int) -> None:
    """Reálný úvěr měsíc po měsíci – vzorci, takže reaguje na změnu úroku."""
    ws = wb.create_sheet("Splátkový kalendář")
    _sirky(ws, {"A": 2, "B": 8, "C": 16, "D": 16, "E": 16, "F": 18})

    ws["B1"] = "Splátkový kalendář (reálný úvěr = CAPEX − vlastní kapitál)"
    ws["B1"].font = NADPIS
    ws["B2"] = "Dopočítává se ze Zadání – po změně úroku nebo délky se přepíše sám."
    ws["B2"].font = SEDY_TEXT

    _hlavicka_tabulky(ws, 4, ["Měsíc", "Splátka", "Úrok", "Úmor", "Zůstatek úvěru"])
    ws["F5"] = "=uver"
    ws["F5"].number_format = KC
    ws["E5"] = "počáteční stav"
    ws["E5"].font = SEDY_TEXT

    for i in range(delka * 12):
        r = 6 + i
        ws.cell(row=r, column=2, value=i + 1)
        for sloupec, vzorec in {
            3: "=splatka_mesic",
            4: f"=$F{r - 1}*urok/12",
            5: f"=$C{r}-$D{r}",
            6: f"=MAX(0,$F{r - 1}-$E{r})",
        }.items():
            c = ws.cell(row=r, column=sloupec, value=vzorec)
            c.number_format = KC_PRESNE
            c.border = RAMECEK

    ws.freeze_panes = "B6"


# ---------------------------------------------------------------- list Úspora zákazníka
def _list_uspora(wb: Workbook, delka: int, prvni_rok: int, radek_samospotreba: int) -> None:
    """Co z kontraktu má klient – proti dnešní vyhnutelné ceně."""
    ws = wb.create_sheet("Úspora zákazníka")
    _sirky(
        ws,
        {"A": 2, "B": 8, "C": 12, "D": 18, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 20},
    )

    ws["B1"] = "Úspora zákazníka"
    ws["B1"].font = NADPIS
    ws["B2"] = (
        "Porovnává se jen vyhnutelná cena (silová složka + vyhnutelné regulované). "
        "Zbytek ceny platí klient tak jako tak, do srovnání nepatří."
    )
    ws["B2"].font = SEDY_TEXT

    _hlavicka_tabulky(
        ws,
        4,
        [
            "Rok",
            "Kalendářní rok",
            "Vyhnutelná cena (Kč/MWh)",
            "Cena PPA (Kč/MWh)",
            "Rozdíl (Kč/MWh)",
            "Samospotřeba (MWh)",
            "Ušetřeno na energii",
            "Nájem baterie",
            "Úspora za rok",
            "Úspora kumulativně",
        ],
    )

    for i in range(delka):
        r = 5 + i
        ws.cell(row=r, column=2, value=i + 1)
        ws.cell(row=r, column=3, value=prvni_rok + i)
        # Kumulativní úspora navazuje na kumulativní sloupec (K), ne na roční (J).
        predchozi = f"$K{r - 1}" if i else "0"
        hodnoty = {
            4: ("=cena_vyhnutelna*(1+index_krok)^INT(($B{r}-1)/index_perioda)", KC),
            5: ("=cena_ppa*(1+index_krok)^INT(($B{r}-1)/index_perioda)", KC),
            6: ("=$D{r}-$E{r}", KC),
            7: (f"=Cashflow!{{s}}{radek_samospotreba}", MWH),
            8: ("=$F{r}*$G{r}", KC),
            9: ("=najem_rok", KC),
            10: ("=$H{r}-$I{r}", KC),
            11: (f"={predchozi}+$J{{r}}", KC),
        }
        for sloupec, (sablona, format_cisla) in hodnoty.items():
            vzorec = sablona.format(r=r, s=get_column_letter(4 + i))
            c = ws.cell(row=r, column=sloupec, value=vzorec)
            c.number_format = format_cisla
            c.border = RAMECEK
        ws.cell(row=r, column=10).font = TUCNE
        ws.cell(row=r, column=11).font = TUCNE

    r = 5 + delka + 1
    ws.cell(row=r, column=2, value="Úspora za celý kontrakt").font = TUCNE
    c = ws.cell(row=r, column=5, value=f"=$K{4 + delka}")
    c.number_format = KC
    c.font = TUCNE

    ws.freeze_panes = "B5"


# ---------------------------------------------------------------- sestavení
def sestav(
    hlavicka: dict,
    varianta: dict,
    vstup: dict,
    p: ParametryEkonomiky,
) -> bytes:
    """Vyrobí sešit z jedné varianty PPA (jedna délka kontraktu).

    `hlavicka` = titulek, podtitulek a datum do záhlaví listu Zadání,
    `varianta` = prvek `po_delkach` z `popis_json`, `vstup` = blok `vstup`
    z `popis_json`, `p` = parametry, se kterými se počítalo.
    """
    delka = int(varianta.get("delka_kontraktu_roky") or 0)
    if delka <= 0:
        raise ValueError("Varianta nemá délku kontraktu, není z čeho sestavit sešit.")

    wb = Workbook()
    # Bez tohohle by IRR, NPV a celý cashflow zůstaly prázdné, dokud v tom někdo
    # ručně nespustí přepočet – openpyxl vzorce zapisuje bez uložených výsledků.
    wb.calculation.fullCalcOnLoad = True

    prvni_rok = int(hlavicka.get("prvni_rok") or datetime.now().year)
    s_baterii = bool(varianta.get("s_baterii"))

    _list_zadani(wb, hlavicka, varianta, vstup, p)
    radky = _list_cashflow(wb, delka, prvni_rok, s_baterii)
    _list_odkup(wb, delka, prvni_rok)
    _list_splatky(wb, delka)
    _list_uspora(wb, delka, prvni_rok, radky["samospotreba"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
