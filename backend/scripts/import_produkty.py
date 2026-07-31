"""Jednorázový import prodejního ceníku z Raynetu do katalogu produktů (CRM-08).

Zdroj: docs/moduly/produkty/Produkty_výběr.xlsx – export „Produkt – Výběr“
z Raynetu (244 položek k 31. 7. 2026).

Spouštět z adresáře backend/ (s aktivním venv):
    python -m scripts.import_produkty                # NÁHLED nasucho, nic nezapíše
    python -m scripts.import_produkty --zapsat       # provede import
    python -m scripts.import_produkty --soubor cesta/k/jinemu.xlsx --zapsat

MAPOVÁNÍ SLOUPCŮ EXCELU:
    Kód               → Technologie.kod (idempotenční klíč)
    Název produktu    → nazev
    Standardní cena   → cena_kc            (prodejní, bez DPH)
    Náklad            → cena_nakup_kc      (nákupní; vidí jen vedení/admin)
    Sazba DPH         → sazba_dph          (0.21 / 0.12 / 0)
    Kategorie         → kategorie          (volný text, i pro filtr)
    Popis             → popis
    Jednotka          → jednotka           („Ks“ i „ks“ se sjednotí na „ks“)
    Platnost od / do  → platnost_od / platnost_do
    Kategorie         → typ (odvozeně, viz `_TYP_PODLE_KATEGORIE`)

CO SE NEIMPORTUJE a proč: Měna (všech 244 řádků je v Kč, appka jinou neumí),
Cena s DPH (dopočet z ceny a sazby – dvě pravdy o jednom čísle), Produktová
řada (u všech 244 prázdná), Zaevidoval / Naposledy změnil (uživatelé Raynetu,
kteří v appce nemusí existovat), Štítky („Import 4.2.2026 #2“ – stopa po
importu do Raynetu, pro nás bezcenná), ID entity (Raynetí ID; kdyby jednou
bylo potřeba, kód `kod` stačí k dohledání).

IDEMPOTENCE: klíčem je `kod`. Existující položku import PŘEPÍŠE jen v cenách,
platnosti a popisu – nesahá na `aktivni` (někdo ji mohl vypnout ručně) ani na
`typ` (mohl ho někdo opravit). Nová položka vznikne se `zdroj="raynet_import"`.
Položky z ceníku BESS (`zdroj="bess_cenik"`) import nikdy nepřepisuje –
poznají se podle zdroje a jejich kW/kWh drží simulaci peak shavingu.
"""

import argparse
from decimal import Decimal, InvalidOperation
from pathlib import Path

# Registrace všech modelů – SQLAlchemy jinak neumí dohledat protějšky vazeb
# (Technologie visí přes položky až na uživatele, objednávky a faktury).
from app.auth import models as _auth_models  # noqa: F401
from app.crm import models as _crm_models  # noqa: F401
from app.finance import models as _finance_models  # noqa: F401
from app.matice import models as _matice_models  # noqa: F401
from app.database import SessionLocal
from app.nabidkovac.models import Technologie

# Kde leží výchozí export z Raynetu (relativně ke kořeni repa).
VYCHOZI_SOUBOR = "docs/moduly/produkty/Produkty_výběr.xlsx"

# Kategorie z Raynetu → typ položky v appce. Typ řídí VÝPOČTY (peak shaving
# bere `typ="baterie"`), kategorie řídí zobrazení. Kategorie, které nejsou
# v mapě, spadnou na „jina“ – správně, protože administrativa ani montážní
# práce do žádného výpočtu nevstupují.
_TYP_PODLE_KATEGORIE = {
    "Panely": "fve_panel",
    "Střídače": "invertor",
    "Baterie": "baterie",
    "BESS": "baterie",
}

# Očekávané pořadí sloupců v exportu (řádek 2 souboru). Kontroluje se před
# importem – kdyby Raynet příště exportoval sloupce jinak, import se zastaví
# místo toho, aby nasypal ceny do špatných polí.
OCEKAVANE_HLAVICKY = [
    "Kód", "Název produktu", "Standardní cena", "Cena s DPH", "Měna", "Sazba DPH",
    "Kategorie", "Popis", "Produktová řada", "Jednotka", "Náklad", "Platnost od",
    "Platnost do", "Zaevidováno", "Zaevidoval", "Naposledy změněno",
    "Naposledy změnil", "Štítky", "ID entity",
]


def _koren_repa() -> Path:
    # backend/scripts/import_produkty.py -> scripts -> backend -> kořen
    return Path(__file__).resolve().parents[2]


def _cislo(hodnota) -> Decimal | None:
    if hodnota is None or hodnota == "":
        return None
    try:
        return Decimal(str(hodnota).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def _datum(hodnota):
    return hodnota.date() if hasattr(hodnota, "date") else hodnota


def nacti_radky(cesta: Path) -> list[dict]:
    """Přečte export a vrátí očištěné řádky. Openpyxl jen tady, ať se kvůli
    jednorázovému skriptu netahá do běhu appky."""
    import openpyxl

    wb = openpyxl.load_workbook(cesta, data_only=True)
    ws = wb.worksheets[0]

    hlavicky = [str(b or "").strip() for b in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    if hlavicky != OCEKAVANE_HLAVICKY:
        chybi = set(OCEKAVANE_HLAVICKY) - set(hlavicky)
        raise SystemExit(
            "Export má jiné sloupce, než skript čeká – import zastaven, ať se ceny "
            f"nenasypou do špatných polí.\nChybí: {sorted(chybi) or '(jen jiné pořadí)'}\n"
            f"Nalezeno: {hlavicky}"
        )

    radky = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        kod = str(r[0]).strip() if r[0] is not None else ""
        nazev = str(r[1]).strip() if r[1] is not None else ""
        if not kod and not nazev:
            continue  # prázdný řádek na konci listu
        kategorie = str(r[6]).strip() if r[6] else ""
        jednotka = str(r[9]).strip().lower() if r[9] else "ks"
        radky.append(
            {
                "kod": kod or None,
                "nazev": nazev or kod,
                "cena_kc": _cislo(r[2]),
                "sazba_dph": _cislo(r[5]),
                "kategorie": kategorie,
                "popis": str(r[7]).strip() if r[7] else "",
                "jednotka": jednotka or "ks",
                "cena_nakup_kc": _cislo(r[10]),
                "platnost_od": _datum(r[11]),
                "platnost_do": _datum(r[12]),
                "typ": _TYP_PODLE_KATEGORIE.get(kategorie, "jina"),
            }
        )
    return radky


def importuj(db, radky: list[dict], zapsat: bool) -> dict:
    """Založí nebo aktualizuje položky. Vrací souhrn pro výpis."""
    podle_kodu = {
        t.kod: t
        for t in db.query(Technologie).filter(Technologie.kod.isnot(None)).all()
    }

    nove, zmenene, preskocene = [], [], []
    for radek in radky:
        stavajici = podle_kodu.get(radek["kod"])
        if stavajici is None:
            nove.append(radek)
            if zapsat:
                db.add(
                    Technologie(
                        typ=radek["typ"],
                        nazev=radek["nazev"],
                        kod=radek["kod"],
                        kategorie=radek["kategorie"],
                        jednotka=radek["jednotka"],
                        popis=radek["popis"],
                        cena_kc=radek["cena_kc"],
                        cena_nakup_kc=radek["cena_nakup_kc"],
                        sazba_dph=radek["sazba_dph"],
                        platnost_od=radek["platnost_od"],
                        platnost_do=radek["platnost_do"],
                        zdroj="raynet_import",
                        aktivni=True,
                    )
                )
            continue

        if stavajici.zdroj == "bess_cenik":
            # Baterie ze simulační matice – import na ně nesahá, jinak by
            # přepsal ceny, na kterých stojí výpočet peak shavingu.
            preskocene.append(radek)
            continue

        zmeny = []
        for pole in ("cena_kc", "cena_nakup_kc", "sazba_dph", "platnost_od", "platnost_do"):
            if getattr(stavajici, pole) != radek[pole]:
                zmeny.append(pole)
                if zapsat:
                    setattr(stavajici, pole, radek[pole])
        for pole in ("nazev", "kategorie", "jednotka", "popis"):
            if (getattr(stavajici, pole) or "") != radek[pole]:
                zmeny.append(pole)
                if zapsat:
                    setattr(stavajici, pole, radek[pole])
        if zmeny:
            zmenene.append((radek, zmeny))

    if zapsat:
        db.commit()

    return {"nove": nove, "zmenene": zmenene, "preskocene": preskocene}


def main():
    parser = argparse.ArgumentParser(
        description="Import prodejního ceníku z Raynetu do katalogu produktů"
    )
    parser.add_argument("--soubor", default=None, help=f"XLSX k importu (výchozí: {VYCHOZI_SOUBOR})")
    parser.add_argument(
        "--zapsat",
        action="store_true",
        help="Bez tohoto přepínače běží NÁHLED nasucho a do DB se nic nezapíše",
    )
    args = parser.parse_args()

    cesta = Path(args.soubor) if args.soubor else _koren_repa() / VYCHOZI_SOUBOR
    if not cesta.exists():
        raise SystemExit(f"Soubor neexistuje: {cesta}")

    radky = nacti_radky(cesta)
    print(f"Načteno {len(radky)} řádků z {cesta.name}")

    kody = [r["kod"] for r in radky if r["kod"]]
    if len(kody) != len(set(kody)):
        raise SystemExit("V exportu jsou duplicitní kódy – import zastaven.")

    db = SessionLocal()
    try:
        vysledek = importuj(db, radky, args.zapsat)
    finally:
        db.close()

    print(f"  nových položek:      {len(vysledek['nove'])}")
    print(f"  aktualizovaných:     {len(vysledek['zmenene'])}")
    print(f"  přeskočených (BESS): {len(vysledek['preskocene'])}")

    if vysledek["nove"][:5]:
        print("\nUkázka nových:")
        for r in vysledek["nove"][:5]:
            print(f"  {r['kod']:<20} {r['nazev'][:55]:<55} {r['cena_kc']} Kč  [{r['kategorie']}]")
    if vysledek["zmenene"][:5]:
        print("\nUkázka změn:")
        for r, zmeny in vysledek["zmenene"][:5]:
            print(f"  {r['kod']:<20} mění se: {', '.join(zmeny)}")

    if not args.zapsat:
        print("\nNÁHLED NASUCHO – nic se nezapsalo. Spusť znovu s --zapsat.")
    else:
        print("\nHotovo, zapsáno do DB.")


if __name__ == "__main__":
    main()
