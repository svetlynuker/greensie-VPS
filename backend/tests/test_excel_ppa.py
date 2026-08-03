"""Testy exportu PPA nabídky do Excelu (`nabidkovac/excel_ppa.py`).

Sešit není snímek čísel, ale živý model – v buňkách jsou vzorce. Riziko proto
není v tom, že by se něco nevypsalo, ale že by to **spočítalo jinak než appka**
a nikdo si toho nevšiml, protože čísla vypadají rozumně. Odtud i to, co se
hlídá:

* **Vzorce a čísla řádků se nesmí rozejít.** Během vývoje se přesně tohle stalo:
  vzorec pro tržbu odkazoval na řádek, kde po přeskládání ležela jiná veličina,
  a sešit tiše počítal s nulou. Proto se řádky adresují klíčem přes
  `RADKY_CASHFLOW` a test kontroluje, že žádný vzorec neodkazuje na jméno,
  které neexistuje.

* **Popisný text nesmí začínat rovnítkem.** Excel takovou buňku bere jako
  vzorec a soubor pak jde otevřít jen s chybou.

* **Sešit počítá totéž co `ppa_v2`.** Tohle ověřuje až přepočet vzorců
  (`formulas`); knihovna je jen pro vývoj, takže se test bez ní přeskočí.
  Spustit ručně: `pip install formulas && pytest tests/test_excel_ppa.py`.

Výběr varianty se schválně nekontroluje tady – jede přes `ppa_tvar`, který má
testy vlastní. Kdyby si Excel vybíral po svém, tiskl by jiný kontrakt než PDF,
které vzniká stejným kliknutím.
"""

import re
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.nabidkovac import excel_ppa, ppa_v2

DELKA = 15
VYROBA_ROK1_MWH = 376.95
PODIL_SS = 0.8
PODIL_SDIL = 0.2
CENA_PPA = 2628.07
KWP = 359.0


def _parametry() -> ppa_v2.ParametryEkonomiky:
    return ppa_v2.parametry_z_nastaveni(None)


def _varianta(s_baterii: bool = False, delka: int = DELKA) -> dict:
    """Varianta v tom tvaru, v jakém ji ukládá `spocti_variantu`.

    Skládá se ze skutečných funkcí `ppa_v2`, ne z ručně opsaných čísel – jinak
    by test potvrzoval jen sám sebe.
    """
    p = _parametry()
    nakladova = KWP * p.nakladova_cena_kc_kwp
    projekt_fve = ppa_v2.sestav_projekt(nakladova, p.marze_fve, p.provize_fve, delka, p)
    baterie_naklad = 800_000.0 if s_baterii else 0.0
    projekt_bess = ppa_v2.sestav_projekt(baterie_naklad, p.marze_bess, p.provize_bess, delka, p)
    najem_mesicni = ppa_v2._najem_baterie_kc_mesic(projekt_bess, p)

    projekt = ppa_v2.Projekt(
        nakladova_cena_kc=projekt_fve.nakladova_cena_kc + projekt_bess.nakladova_cena_kc,
        capex_kc=projekt_fve.capex_kc + projekt_bess.capex_kc,
        provize_kc=projekt_fve.provize_kc + projekt_bess.provize_kc,
        zisk_greensie_kc=projekt_fve.zisk_greensie_kc + projekt_bess.zisk_greensie_kc,
        vlastni_kapital_kc=projekt_fve.vlastni_kapital_kc + projekt_bess.vlastni_kapital_kc,
        uver_kc=projekt_fve.uver_kc + projekt_bess.uver_kc,
        splatka_mesicni_kc=projekt_fve.splatka_mesicni_kc + projekt_bess.splatka_mesicni_kc,
        splatka_rocni_kc=projekt_fve.splatka_rocni_kc + projekt_bess.splatka_rocni_kc,
        delka_roky=delka,
    )
    naklady_rocni = p.servis_kc_rok
    if projekt_bess.capex_kc > 0:
        naklady_rocni += p.bess_servis_kc_rok + p.bess_ems_kc_mesic * 12.0

    cf = ppa_v2.spocti_cashflow(
        VYROBA_ROK1_MWH, PODIL_SS, PODIL_SDIL, CENA_PPA, projekt, p, najem_mesicni * 12.0,
        naklady_rocni,
    )
    return {
        "kwp": KWP,
        "delka_kontraktu_roky": delka,
        "s_baterii": s_baterii,
        "baterie": (
            {
                "kapacita_kwh": 100.0,
                "vykon_kw": 50.0,
                "nakladova_cena_kc": baterie_naklad,
                "najem_kc_mesic": round(najem_mesicni, 2),
            }
            if s_baterii
            else None
        ),
        "cena_ppa_kc_mwh": CENA_PPA,
        "cena_vyhnutelna_kc_mwh": 3460.0,
        "cena_exportu_kc_mwh": p.cena_exportu_kc_mwh,
        "energie": {
            "vyroba_rok1_mwh": VYROBA_ROK1_MWH,
            "samospotreba_mwh": round(VYROBA_ROK1_MWH * PODIL_SS, 3),
            "mira_samospotreby": PODIL_SS,
            "podil_exportu": PODIL_SDIL,
        },
        "financovani": {
            "nakladova_cena_kc": round(projekt.nakladova_cena_kc, 2),
            "capex_kc": round(projekt.capex_kc, 2),
            "provize_kc": round(projekt.provize_kc, 2),
            "zisk_greensie_kc": round(projekt.zisk_greensie_kc, 2),
            "vlastni_kapital_kc": round(projekt.vlastni_kapital_kc, 2),
            "uver_kc": round(projekt.uver_kc, 2),
            "splatka_mesicni_kc": round(projekt.splatka_mesicni_kc, 2),
            "splatka_rocni_kc": round(projekt.splatka_rocni_kc, 2),
            "provozni_naklady_kc_rok": round(naklady_rocni, 2),
        },
        "vysledek_investora": {
            "dscr_min": round(cf.dscr_min, 4) if cf.dscr_min is not None else None,
            "irr": round(cf.irr, 5) if cf.irr is not None else None,
            "npv_kc": round(cf.npv_kc, 2),
        },
        "roky_investor": [
            {
                "rok": r.rok,
                "vyroba_mwh": round(r.vyroba_mwh, 3),
                "samospotreba_mwh": round(r.samospotreba_mwh, 3),
                "sdileni_mwh": round(r.sdileni_mwh, 3),
                "cena_ppa_kc_mwh": round(r.cena_ppa_kc_mwh, 2),
                "prodej_zakaznik_kc": round(r.prodej_zakaznik_kc, 2),
                "prodej_sdileni_kc": round(r.prodej_sdileni_kc, 2),
                "najem_baterie_kc": round(r.najem_baterie_kc, 2),
                "provozni_naklady_kc": round(r.provozni_naklady_kc, 2),
                "zdroje_kc": round(r.zdroje_kc, 2),
                "splatka_kc": round(r.splatka_kc, 2),
                "dscr": round(r.dscr, 4) if r.dscr is not None else None,
                "zisk_po_splatkach_kc": round(r.zisk_po_splatkach_kc, 2),
            }
            for r in cf.roky
        ],
        "odkupni_tabulka": [
            {
                "rok": o.rok,
                "odkupni_cena_kc": round(o.odkupni_cena_kc, 2),
                "zustatek_uveru_kc": round(o.zustatek_uveru_kc, 2),
                "poplatek_predcasne_splaceni_kc": round(o.poplatek_predcasne_splaceni_kc, 2),
                "zisk_spv_kc": round(o.zisk_spv_kc, 2),
            }
            for o in ppa_v2.odkupni_tabulka(projekt, p)
        ],
    }


def _hlavicka() -> dict:
    return {
        "titulek": "PPA výpočet – NAB-26-0007",
        "podtitulek": "Zákazník s.r.o. · FVE 359 kWp",
        "vygenerovano": "Interní model, vygenerováno 3. 8. 2026",
        "prvni_rok": 2027,
    }


def _sesit(**kwargs):
    data = excel_ppa.sestav(_hlavicka(), _varianta(**kwargs), {}, _parametry())
    return load_workbook(BytesIO(data))


# ---- struktura ---------------------------------------------------------------
def test_sesit_ma_vsech_pet_listu():
    wb = _sesit()
    assert wb.sheetnames == [
        "Zadání",
        "Cashflow",
        "Odkup",
        "Splátkový kalendář",
        "Úspora zákazníka",
    ]


def test_delka_jde_podle_kontraktu_ne_napevno():
    """Původní Excel měl 15 a 10 let natvrdo; na kontrakt na 20 let se nedal použít."""
    for delka in (10, 20):
        wb = _sesit(delka=delka)
        assert wb["Cashflow"].cell(row=4, column=3 + delka).value == delka
        assert wb["Cashflow"].cell(row=4, column=4 + delka).value is None
        assert wb["Odkup"].cell(row=4 + delka, column=2).value == delka
        # Splátkový kalendář má řádek na každou měsíční splátku.
        assert wb["Splátkový kalendář"].cell(row=5 + delka * 12, column=2).value == delka * 12


def test_prepocet_pri_otevreni_je_zapnuty():
    """Bez toho zůstanou IRR i celý cashflow prázdné, dokud někdo nestiskne F9."""
    assert _sesit().calculation.fullCalcOnLoad is True


# ---- vzorce ------------------------------------------------------------------
def _vzorce(wb):
    for ws in wb.worksheets:
        for radek in ws.iter_rows():
            for bunka in radek:
                if isinstance(bunka.value, str) and bunka.value.startswith("="):
                    yield ws.title, bunka.coordinate, bunka.value


# Adresa buňky (`D7`, `$L$5`) – hranice kolem, ať se nechytne uvnitř `vyroba_rok1`.
_ADRESA = re.compile(r"(?<![A-Za-zÀ-ž0-9_$])\$?[A-Z]{1,3}\$?\d{1,5}(?![A-Za-zÀ-ž0-9_])")
_LIST = re.compile(r"'[^']+'!|[A-Za-zÀ-ž_][A-Za-zÀ-ž0-9_ ]*!")
_RETEZEC = re.compile(r'"[^"]*"')
_JMENO = re.compile(r"[A-Za-zÀ-ž_][A-Za-zÀ-ž0-9_]*")


def test_zadny_vzorec_neodkazuje_na_neexistujici_nazev():
    """Pojistka proti překlepu v pojmenované buňce – Excel by ukázal #NAME?."""
    wb = _sesit(s_baterii=True)
    zname = set(wb.defined_names)
    for list_nazev, adresa, vzorec in _vzorce(wb):
        holy = _ADRESA.sub(" ", _LIST.sub(" ", _RETEZEC.sub(" ", vzorec)))
        for shoda in _JMENO.finditer(holy):
            slovo = shoda.group()
            if slovo in zname:
                continue
            je_funkce = holy[shoda.end() :].lstrip().startswith("(")
            assert je_funkce, (
                f"{list_nazev}!{adresa}: „{slovo}“ není pojmenovaná buňka ani funkce "
                f"(vzorec {vzorec})"
            )


def test_cashflow_radky_navazuji_klicem_ne_cislem():
    """Tržba musí násobit samospotřebu cenou – ne to, co zrovna leží o řádek výš."""
    wb = _sesit()
    ws = wb["Cashflow"]
    radky = {k: 6 + i for i, (k, *_) in enumerate(excel_ppa.RADKY_CASHFLOW) if k}
    assert ws.cell(row=radky["samospotreba"], column=2).value == "Samospotřeba klienta"
    assert ws.cell(row=radky["cena_ppa_rok"], column=2).value == "Cena PPA"
    prodej = ws.cell(row=radky["prodej_klient"], column=4).value
    assert prodej == f"=D{radky['samospotreba']}*D{radky['cena_ppa_rok']}"
    dscr = ws.cell(row=radky["dscr"], column=4).value
    assert f"D{radky['zdroje']}/D{radky['splatka']}" in dscr


def test_popisky_nezacinaji_rovnitkem():
    """„= výkon × cena" v poznámce Excel bere jako vzorec a soubor nejde otevřít."""
    wb = _sesit()
    for list_nazev, adresa, vzorec in _vzorce(wb):
        assert not re.match(r"^=\s", vzorec), (
            f"{list_nazev}!{adresa} je popisek zapsaný jako vzorec: {vzorec}"
        )


def test_dscr_pod_limitem_se_obarvi():
    ws = _sesit()["Cashflow"]
    pravidla = [p for rozsah in ws.conditional_formatting for p in rozsah.rules]
    assert any(p.operator == "lessThan" and "dscr_limit" in p.formula for p in pravidla)


# ---- vstupy sedí na výsledek výpočtu -----------------------------------------
def test_zlute_vstupy_odpovidaji_vysledku_vypoctu():
    varianta = _varianta(s_baterii=True)
    data = excel_ppa.sestav(_hlavicka(), varianta, {}, _parametry())
    wb = load_workbook(BytesIO(data))

    def bunka(jmeno):
        odkaz = wb.defined_names[jmeno].attr_text
        list_nazev, adresa = odkaz.split("!")
        return wb[list_nazev.strip("'")][adresa.replace("$", "")].value

    assert bunka("kwp") == varianta["kwp"]
    assert bunka("cena_ppa") == varianta["cena_ppa_kc_mwh"]
    assert bunka("roky") == varianta["delka_kontraktu_roky"]
    assert bunka("vyroba_rok1") == varianta["energie"]["vyroba_rok1_mwh"]
    assert bunka("bat_naklad") == varianta["baterie"]["nakladova_cena_kc"]
    assert bunka("urok") == _parametry().urokova_sazba


def test_podil_samospotreby_se_bere_z_mwh_ne_ze_zaokrouhleneho_procenta():
    """Uložené procento má čtyři desetinná místa; na 380 MWh je to rozdíl tisíců Kč."""
    varianta = _varianta()
    varianta["energie"]["mira_samospotreby"] = 0.8002  # zaokrouhleno, jako v popis_json
    presny = varianta["roky_investor"][0]["samospotreba_mwh"] / varianta["roky_investor"][0][
        "vyroba_mwh"
    ]
    data = excel_ppa.sestav(_hlavicka(), varianta, {}, _parametry())
    wb = load_workbook(BytesIO(data))
    odkaz = wb.defined_names["podil_ss"].attr_text
    list_nazev, adresa = odkaz.split("!")
    assert wb[list_nazev.strip("'")][adresa.replace("$", "")].value == pytest.approx(presny)


def test_varianta_bez_delky_je_chyba_ne_prazdny_sesit():
    with pytest.raises(ValueError):
        excel_ppa.sestav(_hlavicka(), {"kwp": 100}, {}, _parametry())


# ---- přepočet vzorců (vyžaduje `formulas`, jinak se přeskočí) ----------------
def _prepocitej(data: bytes, tmp_path):
    formulas = pytest.importorskip(
        "formulas", reason="přepočet vzorců je jen pro vývoj: pip install formulas"
    )
    soubor = tmp_path / "sesit.xlsx"
    soubor.write_bytes(data)
    reseni = formulas.ExcelModel().loads(str(soubor)).finish().calculate()

    def hodnota(list_nazev: str, adresa: str):
        klic = f"]{list_nazev.upper()}'!{adresa.upper()}"
        for k, v in reseni.items():
            if k.upper().endswith(klic):
                return float(v.value[0, 0])
        raise AssertionError(f"buňka {list_nazev}!{adresa} v přepočtu chybí")

    return hodnota


@pytest.mark.parametrize("s_baterii", [False, True])
def test_prepocet_sesitu_dava_stejna_cisla_jako_ppa_v2(tmp_path, s_baterii):
    """Jediný test, který opravdu ověří, že se model v sešitu nerozešel s appkou."""
    varianta = _varianta(s_baterii=s_baterii)
    hodnota = _prepocitej(
        excel_ppa.sestav(_hlavicka(), varianta, {}, _parametry()), tmp_path
    )
    radky = {k: 6 + i for i, (k, *_) in enumerate(excel_ppa.RADKY_CASHFLOW) if k}

    from openpyxl.utils import get_column_letter

    for rok in (1, 2, DELKA):
        ocek = varianta["roky_investor"][rok - 1]
        sl = get_column_letter(3 + rok)
        assert hodnota("Cashflow", f"{sl}{radky['vyroba']}") == pytest.approx(
            ocek["vyroba_mwh"], abs=0.01
        )
        assert hodnota("Cashflow", f"{sl}{radky['prodej_klient']}") == pytest.approx(
            ocek["prodej_zakaznik_kc"], abs=2.0
        )
        assert hodnota("Cashflow", f"{sl}{radky['zdroje']}") == pytest.approx(
            ocek["zdroje_kc"], abs=2.0
        )
        assert hodnota("Cashflow", f"{sl}{radky['dscr']}") == pytest.approx(
            ocek["dscr"], abs=0.001
        )

    radek_cf = 6 + len(excel_ppa.RADKY_CASHFLOW) + 2
    vi = varianta["vysledek_investora"]
    assert hodnota("Cashflow", f"C{radek_cf + 1}") == pytest.approx(vi["irr"], abs=0.0005)
    assert hodnota("Cashflow", f"C{radek_cf + 3}") == pytest.approx(vi["npv_kc"], abs=100.0)
    assert hodnota("Cashflow", f"C{radek_cf + 4}") == pytest.approx(vi["dscr_min"], abs=0.001)

    for rok in (1, DELKA):
        o = varianta["odkupni_tabulka"][rok - 1]
        assert hodnota("Odkup", f"F{4 + rok}") == pytest.approx(o["odkupni_cena_kc"], abs=2.0)
        assert hodnota("Odkup", f"I{4 + rok}") == pytest.approx(o["zisk_spv_kc"], abs=3.0)
        # Splátkový kalendář musí dojít ke stejnému zůstatku jako analytický vzorec.
        assert hodnota("Splátkový kalendář", f"F{5 + rok * 12}") == pytest.approx(
            o["zustatek_uveru_kc"], abs=2.0
        )


# ---- název souboru -----------------------------------------------------------
def test_excel_ma_stejne_jmeno_jako_pdf_jen_jinou_priponu():
    from types import SimpleNamespace

    from app.nabidkovac import pdf as pdf_modul

    nabidka = SimpleNamespace(id=12, cislo="NAB-26-0007")
    kdy = datetime(2026, 8, 3, 23, 55)
    assert pdf_modul.nazev_souboru(nabidka, "ppa", kdy) == "NAB-26-0007_ppa_2026-08-03.pdf"
    assert (
        pdf_modul.nazev_souboru(nabidka, "ppa", kdy, ".xlsx")
        == "NAB-26-0007_ppa_2026-08-03.xlsx"
    )
